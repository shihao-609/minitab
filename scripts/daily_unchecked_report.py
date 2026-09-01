# -*- coding: utf-8 -*-
"""
每日未检验清单自动邮件脚本（独立于 Streamlit 运行）

功能：
  1. 用 service_role 密钥读取全部送检记录 / 检验记录（绕过 RLS）
  2. 复用 inspection_match 比对逻辑，计算「今日未检验清单」
  3. 回算「昨日未检验清单」（送检入库时间 ≤ 昨日 且 质检日期 ≤ 昨日的检验记录），
     对比出新增 / 已解决 / 持续三类变动 —— 无需任何额外存储
  4. 无未检验记录时直接退出（不发邮件）
  5. 按检验工序分组，每个工序一封邮件（Excel 含未检验清单 + 变动摘要 两个 sheet），
     分别发给该工序的收件人/抄送人（收件人可配置「全部」= 所有工序都收）

运行环境变量：
  SUPABASE_URL                Supabase 项目地址
  SUPABASE_SERVICE_ROLE_KEY   服务端密钥（用于绕过 RLS 读全量数据）
  SMTP_HOST                   SMTP 服务器（默认 smtp.qq.com）
  SMTP_PORT                   SMTP 端口（默认 465 / SSL）
  SMTP_USER                   发件邮箱
  SMTP_PASS                   发件邮箱授权码（非登录密码）
  REPORT_RECIPIENTS           收件人邮箱，逗号分隔，可多个
  REPORT_FROM_NAME            发件人显示名（可选，默认"质量管理系统"）
"""
import os
import sys
import smtplib
import time
from datetime import date, datetime, timedelta, timezone
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from io import BytesIO

import pandas as pd
from supabase import create_client

# 让脚本可以 import 项目根目录下的 modules
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from modules import inspection_match  # noqa: E402


def _get_env(name: str, default: str = '') -> str:
    return os.environ.get(name, default) or default


# 每封邮件之间的发送间隔（秒）。凌晨定时任务无需及时，错开发送降低被判定为群发的风险。
SEND_INTERVAL = max(0, int(_get_env('REPORT_SEND_INTERVAL', '30')))
# 单封邮件发送失败后的重试次数（0 = 不重试），及两次重试之间的等待秒数。
MAX_SEND_RETRY = max(0, int(_get_env('REPORT_MAX_RETRY', '2')))
RETRY_INTERVAL = max(0, int(_get_env('REPORT_RETRY_INTERVAL', '60')))


_SRV_CLIENT = None

# 北京时间（与前端发送时间设置一致）
BJ_TZ = timezone(timedelta(hours=8))

_DUMMY_UUID = '00000000-0000-0000-0000-000000000000'


def _load_all_rows(table: str):
    """用 service_role 读取指定表全部数据（绕过 RLS，含所有用户）"""
    global _SRV_CLIENT
    url = _get_env('SUPABASE_URL')
    key = _get_env('SUPABASE_SERVICE_ROLE_KEY')
    if not url or not key:
        raise RuntimeError('缺少 SUPABASE_URL 或 SUPABASE_SERVICE_ROLE_KEY 环境变量')
    if _SRV_CLIENT is None:
        _SRV_CLIENT = create_client(url, key)
    res = _SRV_CLIENT.table(table).select('*').limit(100000).execute()
    return res.data or []


def _load_schedule():
    """读取前端配置的邮件发送排程（report_schedule 表，service_role 绕过 RLS）。

    返回 {'send_days': [1..5], 'send_time': 'HH:MM', 'last_sent': 'YYYY-MM-DD'|''}；
    表已创建但无配置行时自动写入默认行（周一~周五 08:00）；
    表不存在（未创建）时返回 None，脚本回退旧行为（周一~周五发送、无时间门控、无防重复）。
    """
    try:
        rows = _load_all_rows('report_schedule')
    except Exception as e:
        print(f'[排程] 读取 report_schedule 失败（表可能尚未创建），回退默认排程: {e}')
        return None
    if not rows:
        try:
            print('[排程] report_schedule 表为空，写入默认排程（周一~周五 08:00）')
            _SRV_CLIENT.table('report_schedule').delete().neq('id', _DUMMY_UUID).execute()
            _SRV_CLIENT.table('report_schedule').insert(
                {'send_days': '1,2,3,4,5', 'send_time': '08:00'}).execute()
            rows = _load_all_rows('report_schedule')
        except Exception as e:
            print(f'[排程] 初始化默认排程失败: {e}')
            return None
    r = rows[0]
    send_days = sorted({
        int(x) for x in str(r.get('send_days', '')).split(',')
        if x.strip().isdigit() and 1 <= int(x) <= 5
    })
    send_time = str(r.get('send_time') or '').strip()
    last_sent = str(r.get('last_sent_date') or '').strip()
    return {'send_days': send_days, 'send_time': send_time, 'last_sent': last_sent}


def _mark_sent(d: date):
    """记录今日已发送（北京时间），防止同一日重复发送"""
    try:
        _SRV_CLIENT.table('report_schedule').update(
            {'last_sent_date': d.isoformat()}).neq('id', _DUMMY_UUID).execute()
        print(f'[排程] 已记录今日（{d.isoformat()}）发送状态')
    except Exception as e:
        print(f'[排程] 更新最近发送日期失败（不影响本次发送结果）: {e}')


def _check_schedule():
    """按前端配置检查是否应在今天/此刻发送。

    返回 (should_send: bool, reason: str)
    """
    sched = _load_schedule()
    now = datetime.now(BJ_TZ)
    today = now.date()
    if sched is None:
        # 表未创建：回退旧行为（周一~周五发送，无时间门控，无防重复保护）
        if today.weekday() >= 5:
            return False, f'今天是周末（{today}），不在默认周一~周五发送范围内'
        return True, '未配置 report_schedule 表，按默认规则（周一~周五）发送'
    send_days, send_time, last_sent = sched['send_days'], sched['send_time'], sched['last_sent']
    if (today.weekday() + 1) not in send_days:
        return False, f'今天（周{"一二三四五六日"[today.weekday()]}）不在配置的发送日内'
    if send_time and now.strftime('%H:%M') < send_time:
        return False, f'当前 {now.strftime("%H:%M")} 未到配置发送时间 {send_time}'
    if last_sent == today.isoformat():
        return False, f'今日（{today}）已发送过，跳过防重复'
    return True, f'符合排程（周{"一二三四五六日"[today.weekday()]} {send_time or "任意时刻"}）'


def _parse_dt(s):
    """解析 ISO 格式时间（如 '2026-08-21T09:00:00+00:00'），失败返回 None"""
    if not s:
        return None
    try:
        return pd.to_datetime(s)
    except Exception:
        return None


def _today_yesterday():
    """以北京时间为基准计算今天/昨天（与前端发送时间设置一致）"""
    today = datetime.now(BJ_TZ).date()
    return today, today - timedelta(days=1)


def _identity_keys(rows):
    """按 检验类型+五元组 生成身份键（数量规范化与入库 dedup_key 算法一致），支持 DataFrame 或空列表"""
    if rows is None or len(rows) == 0:
        return []
    df = rows if isinstance(rows, pd.DataFrame) else pd.DataFrame(rows)
    t = df['检验类型'].astype(str) if '检验类型' in df.columns else '来料检'
    qty = df['实收数量'].map(inspection_match._fmt_qty) if '实收数量' in df.columns else ''
    return (t + '|' + df['供应商'].fillna('').astype(str) + '|' + df['物料编码'].fillna('').astype(str) + '|'
            + df['规格型号'].fillna('').astype(str) + '|' + df['物料名称'].fillna('').astype(str) + '|'
            + qty).tolist()


def _compare_all_types(sub_df, ins_df):
    """按检验类型分组分别比对再合并，避免跨工序互相匹配（如来料检/过程检）"""
    sub_types = (sorted({str(x) for x in sub_df['检验类型'].astype(str)})
                 if '检验类型' in sub_df.columns and len(sub_df) else ['来料检'])
    ins_types = (sorted({str(x) for x in ins_df['检验类型'].astype(str)})
                 if '检验类型' in ins_df.columns and len(ins_df) else ['来料检'])
    all_types = sorted(set(sub_types) | set(ins_types)) or ['来料检']

    results, checked, unchecked, extra, mismatch_sub, mismatch_ins = [], [], [], [], [], []
    for t in all_types:
        res = inspection_match.compare(sub_df, ins_df, inspect_type=t)
        results.append(res)
        checked.append(res['checked'])
        unchecked.append(res['unchecked'])
        extra.append(res['extra'])
        mismatch_sub.append(res['name_mismatch']['sub'])
        mismatch_ins.append(res['name_mismatch']['ins'])

    def _concat(dfs):
        dfs = [d for d in dfs if d is not None and len(d) > 0]
        return pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame()

    merged = {
        'checked': _concat(checked),
        'unchecked': _concat(unchecked),
        'extra': _concat(extra),
        'name_mismatch': {'sub': _concat(mismatch_sub), 'ins': _concat(mismatch_ins)},
        'summary': {
            'total_sub': int(sum(r['summary']['total_sub'] for r in results)),
            'checked': int(sum(len(r['checked']) for r in results)),
            'unchecked': int(sum(len(r['unchecked']) for r in results)),
            'name_mismatch': int(sum(len(r['name_mismatch']['sub']) for r in results)),
            'extra': int(sum(len(r['extra']) for r in results)),
        },
    }
    return merged


def _load_recipients():
    """读取收件人配置（含适用工序 inspect_type）。
    优先从前端维护的 report_recipients 表读取；表为空或读取失败时回退环境变量 REPORT_RECIPIENTS（视为全部工序）。
    返回: [{'email', 'recipient_type', 'inspect_type'}, ...]
    """
    db_recipients = []
    try:
        for r in _load_all_rows('report_recipients'):
            email = str(r.get('email', '')).strip()
            if not email:
                continue
            db_recipients.append({
                'email': email,
                'recipient_type': 'cc' if str(r.get('recipient_type', '')).strip().lower() == 'cc' else 'to',
                'inspect_type': str(r.get('inspect_type', '全部')).strip() or '全部',
            })
        print(f'[收件人] 从数据库读取配置 {len(db_recipients)} 条（含适用工序）')
    except Exception as e:
        print(f'[收件人] 读取数据库收件人表失败，回退到环境变量: {e}')
    if db_recipients:
        return db_recipients
    env_to = [x.strip() for x in _get_env('REPORT_RECIPIENTS').split(',') if x.strip()]
    return [{'email': e, 'recipient_type': 'to', 'inspect_type': '全部'} for e in env_to]


def build_report():
    today, yesterday = _today_yesterday()
    print(f'[任务] 运行日期: {today}（北京时间），回算基准日: {yesterday}')

    # 0. 发送排程检查（前端可配置周几/几点，仅在此发送）
    should_send, reason = _check_schedule()
    if not should_send:
        print(f'[跳过] {reason}')
        return None
    print(f'[排程] {reason}')

    # 1. 拉取全量数据
    sub_records = _load_all_rows('inspection_submissions')
    ins_records = _load_all_rows('inspection_records')
    print(f'[数据] 送检记录 {len(sub_records)} 条，检验记录 {len(ins_records)} 条')
    if not sub_records:
        print('[跳过] 数据库中无送检记录，不发邮件')
        return None

    sub_df = inspection_match.submissions_to_df(sub_records)
    ins_df = inspection_match.inspection_records_to_df(ins_records)

    # 2. 今日未检验清单（当前全量数据，按检验类型分组比对后合并）
    today_res = _compare_all_types(sub_df, ins_df)
    today_unchecked = today_res['unchecked']
    print(f'[比对] 今日未检验 {len(today_unchecked)} 条，已检验 {len(today_res["checked"])} 条，'
          f'额外检验 {len(today_res["extra"])} 条，名称不一致 {len(today_res["name_mismatch"]["sub"])} 条')

    # 3. 回算昨日未检验清单
    sub_df['_created_dt'] = sub_df['入库时间'].map(_parse_dt)
    ins_df['_inspect_dt'] = pd.to_datetime(ins_df['质检日期'], errors='coerce').dt.date
    yesterday_end = datetime.combine(yesterday, datetime.max.time())

    sub_yesterday = sub_df[sub_df['_created_dt'] <= yesterday_end].drop(columns=['_created_dt'])
    ins_yesterday = ins_df[ins_df['_inspect_dt'] <= yesterday].drop(columns=['_inspect_dt'])
    print(f'[回算] 昨日截止送检 {len(sub_yesterday)} 条，昨日截止检验 {len(ins_yesterday)} 条')

    if len(sub_yesterday) > 0:
        yesterday_res = _compare_all_types(sub_yesterday, ins_yesterday)
        yesterday_unchecked = yesterday_res['unchecked']
    else:
        yesterday_unchecked = []
    print(f'[回算] 昨日未检验 {len(yesterday_unchecked)} 条')

    # 4. 加载收件人配置（含适用工序），失败回退环境变量
    recipients = _load_recipients()
    if not recipients:
        raise RuntimeError('未配置收件人：请在前端「邮件收件人」页面添加收件人，或配置 REPORT_RECIPIENTS 环境变量')
    smtp_user = _get_env('SMTP_USER')
    smtp_pass = _get_env('SMTP_PASS')
    smtp_host = _get_env('SMTP_HOST', 'smtp.qq.com')
    smtp_port = int(_get_env('SMTP_PORT', '465'))
    from_name = _get_env('REPORT_FROM_NAME', '质量管理系统')
    if not smtp_user or not smtp_pass:
        raise RuntimeError('缺少 SMTP_USER 或 SMTP_PASS 环境变量')

    if len(today_unchecked) == 0:
        print('[跳过] 今日无未检验记录，不发邮件')
        return None

    # 5. 按工序分组，每个工序一封邮件，分别发给该工序的收件人/抄送人
    today_unchecked = today_unchecked.reset_index(drop=True)
    yesterday_df = (yesterday_unchecked.reset_index(drop=True)
                    if isinstance(yesterday_unchecked, pd.DataFrame) else pd.DataFrame(yesterday_unchecked))

    sent_emails, total_unchecked, failed_emails = [], 0, []
    for t in sorted({str(x) for x in today_unchecked['检验类型'].astype(str)}):
        t_unchecked = today_unchecked[today_unchecked['检验类型'].astype(str) == t].copy()
        t_keys = set(_identity_keys(t_unchecked))

        # 该工序昨日未检验
        t_yesterday_keys = set()
        if len(yesterday_df) > 0 and '检验类型' in yesterday_df.columns:
            t_yesterday_df = yesterday_df[yesterday_df['检验类型'].astype(str) == t]
            t_yesterday_keys = set(_identity_keys(t_yesterday_df))
        added = t_keys - t_yesterday_keys
        solved = t_yesterday_keys - t_keys
        ongoing = t_keys & t_yesterday_keys

        t_unchecked.insert(0, '状态', [
            '🆕 新增' if k in added else '持续未检验' for k in _identity_keys(t_unchecked)
        ])

        summary_rows = [
            ('检验工序', t),
            ('回算基准日', str(yesterday)),
            ('昨日未检验（该工序）', len(t_yesterday_keys)),
            ('昨日未检验 → 今日已解决', len(solved)),
            ('持续未检验（昨日至今仍未检验）', len(ongoing)),
            ('今日新增未检验', len(added)),
            ('今日未检验合计', len(t_keys)),
        ]

        # 该工序的收件人 / 抄送人（"全部" = 所有工序都收）
        to_list = [r['email'] for r in recipients
                   if r['recipient_type'] == 'to' and r['inspect_type'] in ('全部', t)]
        cc_list = [r['email'] for r in recipients
                   if r['recipient_type'] == 'cc' and r['inspect_type'] in ('全部', t)]
        if not to_list:
            print(f'[跳过] 工序「{t}」有 {len(t_keys)} 条未检验，但未配置收件人（含全部），不发该工序邮件')
            continue

        # 生成该工序的 Excel
        buf = BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            t_unchecked.to_excel(writer, index=False, sheet_name='未检验清单')
            pd.DataFrame(summary_rows, columns=['指标', '数值']).to_excel(
                writer, index=False, sheet_name='变动摘要')

            # 表头加粗 + 红底高亮未检验清单
            from openpyxl.styles import Font, PatternFill
            wb = writer.book
            ws = wb['未检验清单']
            for cell in ws[1]:
                cell.font = Font(bold=True)
            if len(t_unchecked) > 0:
                red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                for row in ws.iter_rows(min_row=2, max_row=len(t_unchecked) + 1,
                                        min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.fill = red

        filename = f'未检验清单_{t}_{today.isoformat()}.xlsx'
        print(f'[输出] 工序「{t}」生成 {filename}，未检验 {len(t_keys)} 条')

        # 发送：使用纯邮箱地址作为 From，避免中文显示名触发 SMTPDataError 导致重复投递
        smtp_sender = str(smtp_user).strip().strip('"').strip("'")
        display_name = str(from_name or '').strip().strip('"').strip("'")
        msg = MIMEMultipart()
        msg['From'] = smtp_sender
        msg['To'] = ', '.join(to_list)
        if cc_list:
            msg['Cc'] = ', '.join(cc_list)
        msg['Subject'] = f'【未检验清单·{t}】{today.isoformat()} 共 {len(t_keys)} 条'
        body = f'「{t}」未检验清单详见附件。\n\n{display_name}' if display_name else f'「{t}」未检验清单详见附件。'
        msg.attach(MIMEText(body, 'plain', 'utf-8'))

        part = MIMEApplication(buf.getvalue(), _subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        part.add_header('Content-Disposition', 'attachment', filename=('utf-8', '', filename))
        msg.attach(part)

        smtp_recipients = to_list + cc_list
        last_err = None
        for attempt in range(1, MAX_SEND_RETRY + 2):
            try:
                print(f'[邮件] 工序「{t}」连接 {smtp_host}:{smtp_port}（第 {attempt} 次尝试）...')
                with smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=30) as server:
                    server.login(smtp_user, smtp_pass)
                    server.sendmail(smtp_sender, smtp_recipients, msg.as_string())
                last_err = None
                break
            except Exception as e:
                last_err = e
                print(f'[邮件] 工序「{t}」第 {attempt} 次发送失败: {e}')
                if attempt <= MAX_SEND_RETRY:
                    print(f'[等待] {RETRY_INTERVAL} 秒后重试...')
                    time.sleep(RETRY_INTERVAL)
        if last_err is not None:
            print(f'[失败] 工序「{t}」重试 {MAX_SEND_RETRY} 次后仍发送失败，跳过该工序继续后续: {last_err}')
            failed_emails.append(filename)
            if SEND_INTERVAL > 0:
                time.sleep(SEND_INTERVAL)
            continue
        cc_txt = f'，抄送 {len(cc_list)} 人' if cc_list else ''
        print(f'[邮件] 工序「{t}」已发送给 {len(to_list)} 个收件人{cc_txt}: {", ".join(smtp_recipients)}')
        sent_emails.append(filename)
        total_unchecked += len(t_keys)
        if SEND_INTERVAL > 0:
            print(f'[等待] 下一封邮件间隔 {SEND_INTERVAL} 秒（凌晨任务，错开发送）')
            time.sleep(SEND_INTERVAL)

    if not sent_emails:
        print('[完成] 所有工序均未发送（未配置对应收件人，或没有未检验记录）')
        return None
    print(f'[完成] 共发送 {len(sent_emails)} 封邮件，合计未检验 {total_unchecked} 条')
    if failed_emails:
        print(f'[警告] {len(failed_emails)} 封发送失败: {", ".join(failed_emails)}')
    # 记录今日已发送，防止 hourly 触发下同一日重复发送
    _mark_sent(today)
    return {'emails': sent_emails, 'unchecked_count': total_unchecked, 'failed': failed_emails}


def main():
    try:
        result = build_report()
        if result:
            print(f'[完成] 共发送 {len(result["emails"])} 封邮件（未检验合计 {result["unchecked_count"]} 条）')
            if result.get('failed'):
                print(f'[警告] 其中 {len(result["failed"])} 封发送失败: {", ".join(result["failed"])}')
        else:
            print('[完成] 无需发送（未检验清单为空或没有送检记录）')
    except Exception as e:
        print(f'[失败] {e}', file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
