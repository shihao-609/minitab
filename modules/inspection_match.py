"""
送检清单 vs 检验清单 对比模块
================================
功能：
  1. 解析送检清单 Excel（6 核心列：供应商 / 物料编码 / 规格型号 / 物料名称 / 收料日期 / 实收数量）
  2. 解析检验清单 Excel（6 核心列：供应商 / 物料编码 / 规格型号 / 物料名称 / 质检日期 / 检验数量）
  3. 送检记录幂等去重入库（去重键：供应商+物料编码+规格型号+物料名称+实收数量，日期不参与）
  4. 对比分类：✅ 已检验 / ⚠️ 未检验 / 📋 额外检验 / 🆔 名称不一致
  5. Excel 模板与结果导出

支持的检验清单字段（2026-08-21 用户提供）：
  是否加急检验、单据编号、数据状态、质检日期、供应商、采购员、批号、质检员、
  物料编码、物料名称、规格型号、单位、检验数量、合格数、不合格数、检验结果、
  异常描述、备注、类别、不良原因分析、解决措施、进度跟踪、不合格处理单号、
  检验工时、二次合格率、归属部件、归属产品、二次异常描述、检验方法、用途

设计要点：
  - 逐批次严格对比：每条送检记录需找到五字段完全一致的检验记录（质检日期 ≥ 收料日期）
  - 检验数量必须 = 实收数量才视为匹配
  - 检验清单不入库，仅用于本次临时对比；非核心列自动忽略
  - 名称不一致单独成类，人工判断
"""
import hashlib
import re
from datetime import date, datetime, timedelta
from io import BytesIO

import pandas as pd

from modules import supabase_helper
from modules.receipt_notice import fill_receipt_order
from modules.supplier_normalize import normalize_supplier, same_supplier

SUBMISSION_COLS = ['供应商', '物料编码', '规格型号', '物料名称', '收料日期', '实收数量']
INSPECTION_COLS = ['供应商', '物料编码', '规格型号', '物料名称', '质检日期', '检验数量']

# ==================== 检验工序配置中心 ====================
# 以后新增检验工序（如过程检/出货检）时，只需在下方加一项配置：
#   - match_pairs: 该工序「送检清单 ↔ 检验清单」用于唯一匹配的字段对。
#                  **不同工序字段可以完全不同**（如来料检按供应商+物料，过程检可能按工单号+批次号）。
#                  - 比对（compare）用它做匹配键与差异核对
#                  - 入库去重用它计算通用去重键 dedup_key（数据库唯一索引不绑定固定列）
#   - ins_dedup_extra: 检验清单去重额外字段（默认单据编号+质检日期），送检清单无需额外字段
#   - 前端 app.py 的 INSPECT_TYPES 列表同步加一项，上传模板/Excel 增加对应列即可
INSPECT_TYPE_CONFIGS = {
    '来料检': {
        'match_pairs': [
            ('供应商', '供应商'),
            ('物料编码', '物料编码'),
            ('规格型号', '规格型号'),
            ('物料名称', '物料名称'),
            ('实收数量', '检验数量'),
        ],
        'ins_dedup_extra': ['单据编号', '质检日期'],
    },
    # 示例（以后加过程检时取消注释并按需修改字段）：
    # '过程检': {
    #     'match_pairs': [
    #         ('工单号', '工单号'),
    #         ('物料编码', '物料编码'),
    #         ('工序名称', '工序名称'),
    #         ('批次号', '批次号'),
    #         ('送检数量', '检验数量'),
    #     ],
    #     'ins_dedup_extra': ['单据编号', '质检日期'],
    # },
}


def get_inspect_type_config(inspect_type='来料检'):
    """获取某工序的匹配配置（未配置的工序回退到来料检）"""
    return INSPECT_TYPE_CONFIGS.get(inspect_type or '来料检', INSPECT_TYPE_CONFIGS['来料检'])

# 检验清单入库时保留的扩展字段（非必填，缺失时置空）
EXTRA_INS_COLS = {
    '单据编号': ['单据编号', '检验单号', '单号', '检验报告编号'],
    '合格数': ['合格数', '合格数量', '合格'],
    '不合格数': ['不合格数', '不合格数量', '不合格', '不良数'],
    '检验结果': ['检验结果', '检验结论', '判定', '结果'],
    '质检员': ['质检员', '检验员', '检验人', '质检人'],
    '批号': ['批号', '批次', '批号/序号'],
    '类别': ['类别', '检验类别', '检验方式', '分类'],
    '采购员': ['采购员', '采购人', 'buyer'],
}

# 送检清单可选展示列（非匹配/去重字段，仅用于展示，缺失时置空）
SUBMISSION_EXTRA_COLS = ['采购员']

COLUMN_ALIASES = {
    '供应商': ['供应商', 'supplier', '供应商名称', 'vendor'],
    '物料编码': ['物料编码', '编码', '料号', '物料代码', '物料号', 'material_code', 'code'],
    '规格型号': ['规格型号', '规格', '型号', 'spec', 'specification'],
    '物料名称': ['物料名称', '名称', '品名', '物料名', 'material', 'name'],
    '收料日期': ['收料日期', '收货日期', '收料时间', '到货日期', '来料日期'],
    '实收数量': ['实收数量', '实收数', '实收量', '实收', '数量'],
    '质检日期': ['质检日期', '检验日期', '检验时间', '质检时间', '检验完成日期', '质检完成日期'],
    '检验数量': ['检验数量', '检验数', '检验量', '检验'],
}

DATE_FORMATS = ['%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S', '%Y-%m-%d %H:%M',
                '%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y%m%d', '%Y年%m月%d日']


# ==================== 基础清洗 ====================

def _clean_text(v):
    if v is None:
        return ''
    try:
        if pd.isna(v):
            return ''
    except Exception:
        pass
    return str(v).strip()


def _normalize_header(c):
    """
    列名规范化：去掉首尾空格、尾部星号、尾部的括号说明（如 "检验结果*"、"类别（下拉框）"）。
    用于兼容 ERP 导出里带星号或括号注释的表头。
    """
    s = _clean_text(c)
    s = re.sub(r'\*+\s*$', '', s)
    s = re.sub(r'[（(].*?[）)]\s*$', '', s).strip()
    return s


def _normalize_code(v):
    """物料编码规范化：数值去除 .0，字符串去空格"""
    if v is None:
        return ''
    try:
        if pd.isna(v):
            return ''
    except Exception:
        pass
    if isinstance(v, bool):
        return ''
    if isinstance(v, (int, float)):
        if float(v).is_integer():
            return str(int(v))
        return str(v)
    s = str(v).strip()
    if re.fullmatch(r'\d+\.0', s):
        return s[:-2]
    return s


def _parse_qty(v):
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except Exception:
        pass
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(',', '').replace('，', '').replace(' ', '')
    m = re.sub(r'[^\d.+-]', '', s)
    if m in ('', '.', '-', '+'):
        return None
    try:
        return float(m)
    except Exception:
        return None


def _parse_date(v, default=None):
    if v is None:
        return default
    try:
        if pd.isna(v):
            return default
    except Exception:
        pass
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, pd.Timestamp):
        return v.date()
    if isinstance(v, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(v))).date()
        except Exception:
            return default
    s = str(v).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    s2 = re.split(r'[\sT]', s)[0]
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s2, fmt).date()
        except Exception:
            continue
    return default


def _fmt_date(v):
    d = _parse_date(v)
    return d.isoformat() if d else ''


def _fmt_qty(v):
    v = _parse_qty(v)
    if v is None:
        return ''
    if v == int(v):
        return str(int(v))
    return str(v)


# ==================== Excel 解析 ====================

def _find_header_row(df):
    """在前 15 行中寻找包含物料编码/物料名称等关键列名的表头行"""
    for idx in range(min(15, len(df))):
        row = df.iloc[idx].astype(str)
        hits = 0
        for col in row:
            c = _normalize_header(col).lower()
            if any(c == a.lower() for a in COLUMN_ALIASES['物料编码']):
                hits += 1
            elif any(c == a.lower() for a in COLUMN_ALIASES['物料名称']):
                hits += 1
            elif any(c == a.lower() for a in COLUMN_ALIASES['供应商']):
                hits += 1
        if hits >= 2:
            return idx
    return None


def _map_columns(df, kind):
    target = SUBMISSION_COLS if kind == 'submission' else INSPECTION_COLS
    remap = {}
    used = set()
    df_cols = [_normalize_header(c) for c in df.columns]
    for std in target:
        for i, col in enumerate(df_cols):
            if i in used:
                continue
            if any(a.lower() == col.lower() for a in COLUMN_ALIASES[std]):
                remap[df.columns[i]] = std
                used.add(i)
                break
    missing = [c for c in target if c not in remap.values()]
    return df.rename(columns=remap), missing


def parse_sheet(uploaded_file, kind='submission', default_date=None, clean_report=None):
    """
    解析上传的 Excel，返回标准化后的 DataFrame（6 标准列）。

    收料日期（kind='submission'）：
      - 若识别为 ERP 收料通知单（含「单据编号」列）→ 自动按单据号向下填充
        收料日期/单据状态/供应商/整单关闭状态（ERP 导出时可能仅首行有值），
        填充明细写入 clean_report（若传入 dict）；「整单关闭状态 = 单据关闭」的行
        表示该单已检验完毕，会被跳过（计数写入 clean_report['closed_rows']）；
      - 普通 6 列送检清单 → 按 Excel 原样保存，为空保持为空。
    参数 default_date 仅作向后兼容保留，已不使用。
    """
    df = pd.read_excel(uploaded_file, header=None)
    header_row = _find_header_row(df)
    if header_row is None:
        raise ValueError('未找到表头行，请确保包含「物料编码」「物料名称」等列名')
    df.columns = df.iloc[header_row].astype(str).str.strip()
    df = df.iloc[header_row + 1:].reset_index(drop=True)
    df = df.dropna(how='all')
    # 识别 ERP 收料通知单（含「单据编号」列）：自动按单据号填充收料日期/单据状态/供应商/整单关闭状态
    if kind == 'submission':
        norm = {_normalize_header(c): c for c in df.columns}
        rename_map = {}
        for _key in ('单据编号', '供应商', '整单关闭状态', '采购员'):
            if _key in norm and norm[_key] != _key:
                rename_map[norm[_key]] = _key
        if rename_map:
            df = df.rename(columns=rename_map)
        if '单据编号' in df.columns:
            df, clean_notes = fill_receipt_order(df)
            if isinstance(clean_report, dict):
                clean_report.update(clean_notes)
    # 整单关闭状态 =「单据关闭」表示该单据已检验完毕，不参与送检/检验比对 → 解析时跳过
    if kind == 'submission' and '整单关闭状态' in df.columns:
        _closed = df['整单关闭状态'].astype(str).str.strip().eq('单据关闭')
        if _closed.any():
            if isinstance(clean_report, dict):
                clean_report['closed_rows'] = int(_closed.sum())
            df = df[~_closed].reset_index(drop=True)
    df, missing = _map_columns(df, kind)
    if missing:
        raise ValueError(f'缺少必要列: {", ".join(missing)}（请使用标准列名或下载模板）')

    # 可选列：采购员（送检清单展示用，缺失置空；不参与匹配/去重）
    if kind == 'submission':
        _purchaser_col = None
        for _c in df.columns:
            if _normalize_header(_c).lower() in ('采购员', '采购人', 'buyer'):
                _purchaser_col = _c
                break
        df['采购员'] = df[_purchaser_col].map(_clean_text) if _purchaser_col else ''

    df['供应商'] = df['供应商'].map(_clean_text)
    # 过滤 Excel 底部的「合计」汇总行（供应商列含"合计"，如 ERP 导出常见的合计行）
    df = df[~df['供应商'].str.contains('合计', na=False)].reset_index(drop=True)
    df['物料编码'] = df['物料编码'].map(_normalize_code)
    df['规格型号'] = df['规格型号'].map(_clean_text)
    df['物料名称'] = df['物料名称'].map(_clean_text)
    if kind == 'submission':
        df['收料日期'] = df['收料日期'].map(lambda v: _parse_date(v, default=None))
        df['实收数量'] = df['实收数量'].map(_parse_qty)
    else:
        df['质检日期'] = df['质检日期'].map(lambda v: _parse_date(v, default=None))
        df['检验数量'] = df['检验数量'].map(_parse_qty)

    df = df[df['物料编码'] != ''].reset_index(drop=True)
    if kind == 'submission':
        return df[SUBMISSION_COLS + SUBMISSION_EXTRA_COLS]
    return df[INSPECTION_COLS]


# ==================== 幂等去重入库 ====================

def _dedup_val(field, v):
    """把某字段值规范化为「去重键用」的稳定文本（与数据库回填 SQL 的算法一致）"""
    if field in ('实收数量', '检验数量', '合格数', '不合格数'):
        return _fmt_qty(_parse_qty(v))
    if field in ('收料日期', '质检日期'):
        d = _parse_date(v)
        return d.isoformat() if d else ''
    if field == '物料编码':
        return _normalize_code(v)
    return _clean_text(v)


def make_dedup_key(row, inspect_type='来料检', kind='submission'):
    """
    按工序配置计算通用去重键（MD5）。入库时写入表的 dedup_key 列，
    数据库唯一索引统一为 (inspect_type, dedup_key)（团队共享全局去重），与字段多少无关。

    row: 字典，键为中文列名（供应商/物料编码/...）。
    kind: 'submission' 用 match_pairs 送检侧字段；'inspection' 再加 ins_dedup_extra。
    """
    cfg = get_inspect_type_config(inspect_type)
    fields = [sc for sc, _ in cfg['match_pairs']]
    if kind == 'inspection':
        fields = [ic for _, ic in cfg['match_pairs']] + list(cfg.get('ins_dedup_extra', []))
    raw = '|'.join(_dedup_val(f, row.get(f)) for f in fields)
    return hashlib.md5(raw.encode('utf-8')).hexdigest()


def _existing_keys(records, kind='submission'):
    """已入库记录的 (检验类型, dedup_key) 集合"""
    # 数据库返回的键是英文列名，兜底计算时需转回中文列名（与 make_dedup_key 的输入一致）
    _EN_CN = {
        'supplier': '供应商', 'material_code': '物料编码', 'spec': '规格型号',
        'material_name': '物料名称', 'received_date': '收料日期', 'received_qty': '实收数量',
        'inspect_date': '质检日期', 'inspect_qty': '检验数量', 'doc_no': '单据编号',
        'batch_no': '批号',
    }
    keys = set()
    for r in records:
        it = _clean_text(r.get('inspect_type')) or '来料检'
        dk = _clean_text(r.get('dedup_key'))
        if dk:
            keys.add((it, dk))
        else:
            # 老数据兜底：英文键转中文后按工序配置计算（与入库算法一致）
            row_cn = {cn: r.get(en) for en, cn in _EN_CN.items()}
            keys.add((it, make_dedup_key(row_cn, it, kind)))
    return keys


def preview_import(df, records=None, inspect_type='来料检'):
    """
    返回 (新增记录 df, 重复记录 df)。去重键 = 检验类型 + 通用 dedup_key（按工序配置字段计算）。

    records 为空时走轻量查询（只拉取键列），避免全量 SELECT *。
    """
    if records is None:
        records = supabase_helper.fetch_submission_keys()
    keys = _existing_keys(records)

    df = df.copy()
    df['检验类型'] = inspect_type
    df['去重键'] = df.apply(lambda r: make_dedup_key(r, inspect_type, 'submission'), axis=1)

    key_series = df['检验类型'] + '|' + df['去重键']
    is_new = ~key_series.isin({f'{it}|{dk}' for it, dk in keys})
    return df[is_new].reset_index(drop=True), df[~is_new].reset_index(drop=True)


def import_submissions(df, progress=None, batch_size=1000, inspect_type='来料检'):
    """
    幂等入库：收料日期保持 Excel 原样（为空则入库为空）；类型+五元组重复的跳过。
    返回 (插入数, 重复数, 新增行数)

    progress: 可选回调 progress(done, total)
    """
    df = df.copy()

    # 直接插入全部数据，由数据库 unique index (inspect_type, dedup_key) 完成去重（团队共享全局去重）
    total = len(df)
    rows = []
    for _, row in df.iterrows():
        d = row['收料日期']
        rows.append({
            'supplier': row['供应商'],
            'material_code': row['物料编码'],
            'spec': row['规格型号'],
            'material_name': row['物料名称'],
            'received_date': d.isoformat() if isinstance(d, (date, datetime)) else None,
            'received_qty': row['实收数量'],
            'purchaser': _clean_text(row.get('采购员', '')),
            'inspect_type': inspect_type,
            'dedup_key': make_dedup_key(row, inspect_type, 'submission'),
        })

    inserted = 0
    if rows:
        for i in range(0, len(rows), batch_size):
            batch = rows[i:i + batch_size]
            inserted += supabase_helper.insert_inspection_submissions(batch)
            if progress:
                progress(min(i + len(batch), total), total)

    skipped = total - inserted
    return inserted, skipped, total


def submissions_to_df(records):
    """数据库记录 → 展示用 DataFrame（检验类型 + 6 标准列 + 采购员 + 入库时间 + id + 去重键）"""
    rows = [{
        '检验类型': _clean_text(r.get('inspect_type')) or '来料检',
        '供应商': _clean_text(r.get('supplier')),
        '物料编码': _normalize_code(r.get('material_code')),
        '规格型号': _clean_text(r.get('spec')),
        '物料名称': _clean_text(r.get('material_name')),
        '收料日期': _fmt_date(r.get('received_date')),
        '实收数量': _parse_qty(r.get('received_qty')),
        '采购员': _clean_text(r.get('purchaser')),
        '入库时间': str(r.get('created_at', ''))[:19],
        'id': r.get('id'),
        '去重键': _clean_text(r.get('dedup_key')),
    } for r in records]
    return pd.DataFrame(rows, columns=['检验类型', '供应商', '物料编码', '规格型号', '物料名称',
                                       '收料日期', '实收数量', '采购员', '入库时间', 'id', '去重键'])


# ==================== 对比 ====================

def _qty_eq(a, b):
    a, b = _parse_qty(a), _parse_qty(b)
    if a is None or b is None:
        return False
    return abs(a - b) < 1e-6


def _date_ge(ins_date, sub_date):
    """质检日期 >= 收料日期；任一为空视为通过"""
    if ins_date is None or sub_date is None:
        return True
    d1 = _parse_date(ins_date)
    d2 = _parse_date(sub_date)
    if d1 is None or d2 is None:
        return True
    return d1 >= d2


def _sub_dict(row):
    d = {c: row.get(c, '') for c in SUBMISSION_COLS}
    d['采购员'] = row.get('采购员', '')
    d['检验类型'] = row.get('检验类型', '来料检')
    return d


def _ins_dict(row):
    d = {c: row.get(c, '') for c in INSPECTION_COLS}
    d['采购员'] = row.get('采购员', '')
    d['检验类型'] = row.get('检验类型', '来料检')
    return d


_LABELS = {'供应商': '供应商', '规格型号': '规格', '物料名称': '名称', '物料编码': '编码'}


def _norm_cached(name, cache):
    """带缓存的 normalize_supplier：同一名称只归一化一次。
    normalize_supplier 是纯函数，缓存只影响耗时，不影响结果。"""
    if name in cache:
        return cache[name]
    r = normalize_supplier(name)
    cache[name] = r
    return r


def _same_supplier_cached(a, b, cache):
    """与 supplier_normalize.same_supplier 语义逐字一致，仅复用归一化缓存。
    a/b 为 _fill_row 已规范化（_clean_text）后的字符串，_clean_text 与 _clean 对已规范值等价，
    因此判定结果与原 same_supplier 完全相同。"""
    a = _clean_text(a)
    b = _clean_text(b)
    if not a or not b:
        return a == b
    if a == b:
        return True
    na = _norm_cached(a, cache)
    nb = _norm_cached(b, cache)
    if not na or not nb:
        return False
    if na == nb:
        return True
    short, long = (na, nb) if len(na) <= len(nb) else (nb, na)
    if len(short) >= 2 and short in long and len(long) - len(short) <= 8:
        return True
    return False


def _quick_fields_equal(sub, ins, pairs):
    """快速必然性检查：非供应商字段全部精确相等才可能匹配。

    匹配成功要求 _diff_fields 全字段 diff 为空，其中非供应商字段都是精确比较；
    因此本函数返回 False 时，完整 _diff_fields 必非空，该候选不可能匹配成功。
    仅用于提前排除候选，不生成差异文本、不调用供应商归一化。
    """
    for sc, ic in pairs:
        if sc == '供应商':
            continue
        sv, iv = sub.get(sc), ins.get(sc)
        if sc in ('实收数量', '检验数量'):
            if not _qty_eq(sv, iv):
                return False
        elif _clean_text(sv) != _clean_text(iv):
            return False
    return True


def _diff_fields(sub, ins, cfg, use_norm_supplier=True, sup_cache=None):
    """严格核对送检记录与检验记录的字段，返回差异描述（空串表示完全一致）。

    核对字段来自工序配置 INSPECT_TYPE_CONFIGS 的 match_pairs（各工序可不同）。
    use_norm_supplier=True 时供应商用归一化比较（自动识别同一公司的不同写法）。
    sup_cache: 可选 dict，缓存供应商归一化结果，加速大量记录的重复比较。
    """
    diffs = []
    for sc, ic in cfg['match_pairs']:
        sv, iv = sub.get(sc), ins.get(sc)  # ins 记录已统一用送检侧字段名存储
        label = _LABELS.get(sc, sc if sc == ic else f'{sc}/{ic}')
        if sc == '供应商':
            if use_norm_supplier:
                ok = _same_supplier_cached(sv, iv, sup_cache) if sup_cache is not None else same_supplier(sv, iv)
            else:
                ok = _clean_text(sv) == _clean_text(iv)
            if not ok:
                diffs.append(f"供应商「{sv or '空'}」≠「{iv or '空'}」")
            continue
        if sc in ('实收数量', '检验数量'):
            if not _qty_eq(sv, iv):
                diffs.append(f"{_LABELS.get(sc, sc)} {_fmt_qty(sv) or '空'} ≠ {_fmt_qty(iv) or '空'}")
            continue
        if _clean_text(sv) != _clean_text(iv):
            diffs.append(f"{label}「{sv or '空'}」≠「{iv or '空'}」")
    return '；'.join(diffs)


def compare(sub_df, ins_df, progress=None, normalize_suppliers=True, inspect_type=None):
    """
    核心对比。逐批次严格匹配，返回四分类结果 dict：
      checked           ✅ 已检验（附匹配的检验记录摘要）
      unchecked         ⚠️ 未检验
      extra             📋 额外检验（送检清单中不存在的检验记录）
      name_mismatch     🆔 名称不一致（同编码但物料名称不同，需人工判断）
      summary           各分类计数

    normalize_suppliers: 开启后自动识别同一供应商的不同写法（默认开启）。
    inspect_type: 指定检验类型（如 '来料检'）时只比对该类型的送检/检验记录，
                  避免跨工序互相匹配；None 表示全量比对。
    progress: 可选回调 progress(done, total)
    """
    sub_df = sub_df.reset_index(drop=True)
    ins_df = ins_df.reset_index(drop=True)

    if inspect_type:
        if '检验类型' in sub_df.columns:
            sub_df = sub_df[sub_df['检验类型'] == inspect_type].reset_index(drop=True)
        if '检验类型' in ins_df.columns:
            ins_df = ins_df[ins_df['检验类型'] == inspect_type].reset_index(drop=True)

    def _norm_sup(name):
        return normalize_supplier(name) if normalize_suppliers else _clean_text(name)

    # 按工序配置取匹配字段（各工序可不同）
    cfg = get_inspect_type_config(inspect_type or '来料检')
    pairs = cfg['match_pairs']
    sub_fields = [sc for sc, _ in pairs]
    has_name = '物料名称' in sub_fields
    has_code = '物料编码' in sub_fields
    sub_date_field = cfg.get('sub_date_field', '收料日期')
    ins_date_field = cfg.get('ins_date_field', '质检日期')
    has_date = (sub_date_field in sub_df.columns and ins_date_field in ins_df.columns)

    # 预处理：to_dict 批量转换（比 iterrows 快约 3 倍）
    # 送检/检验两侧统一用「送检侧字段名」存值，方便按配置生成匹配键
    def _fill_row(r, is_ins):
        row = {'检验类型': _clean_text(r.get('检验类型', '来料检')) or '来料检'}
        for sc, ic in pairs:
            src = ic if is_ins else sc
            v = r.get(src)
            if sc == '供应商':
                sup = _clean_text(v)
                row[sc] = sup
                row['_sup_norm'] = _norm_sup(sup)
            elif sc == '物料编码':
                row[sc] = _normalize_code(v)
            elif sc in ('实收数量', '检验数量'):
                row[sc] = _parse_qty(v)
            else:
                row[sc] = _clean_text(v)
        row['采购员'] = _clean_text(r.get('采购员', ''))  # 展示用，不参与匹配/去重
        if has_date:
            src_date = ins_date_field if is_ins else sub_date_field
            row[src_date] = _parse_date(r.get(src_date))
        return row

    subs = [_fill_row(r, False) for r in sub_df.to_dict('records')]
    ins_rows = [_fill_row(r, True) for r in ins_df.to_dict('records')]

    # 建立索引（供应商键使用归一化名；匹配键 = 工序配置的字段元组）
    ins_by_code = {}
    ins_by_full_key = {}
    code_name_sets = {}
    for idx, c in enumerate(ins_rows):
        if has_code:
            code = c['物料编码']
            ins_by_code.setdefault(code, []).append((idx, c))
            if has_name:
                code_name_sets.setdefault(code, set()).add(c['物料名称'])
        key = tuple(c[f] for f in sub_fields)
        ins_by_full_key.setdefault(key, []).append((idx, c))

    def _sub_match_key(row):
        return tuple(row[f] for f in sub_fields)

    checked, unchecked = [], []
    name_mismatch_sub, name_mismatch_ins = [], []
    matched_ins = set()

    _sup_cache = {}  # 供应商归一化缓存（纯加速，判定结果不变）
    subs_by_code = {}
    if has_code:
        for s in subs:
            subs_by_code.setdefault(s['物料编码'], []).append(s)

    total_sub = len(subs)
    report_every = max(1, total_sub // 20)  # 每 5% 进度报告一次，避免频繁重绘

    for s_idx, srow in enumerate(subs):
        code = srow['物料编码'] if has_code else None
        cands_code = ins_by_code.get(code, []) if has_code else []

        if not cands_code:
            unchecked.append(_sub_dict(srow))
            continue

        # 先尝试全键命中（O(1)），匹配键 = 工序配置的字段元组
        full_key = _sub_match_key(srow)
        candidates = ins_by_full_key.get(full_key, [])
        matched_idx = None
        matched_ins_row = None
        for idx, c in candidates:
            if idx in matched_ins:
                continue
            if not has_date or _date_ge(c[ins_date_field], srow[sub_date_field]):
                matched_idx = idx
                matched_ins_row = c
                break

        if matched_idx is not None:
            d = _sub_dict(srow)
            _mr = matched_ins_row or {}
            d['匹配检验记录'] = (f"{_mr.get('供应商', '') or ''} · {_mr.get('规格型号', '') or ''} · "
                              f"检验数量 {_fmt_qty(_mr.get('检验数量'))} · "
                              f"质检日期 {_fmt_date(_mr.get('质检日期'))}")
            checked.append(d)
            matched_ins.add(matched_idx)
        else:
            # 名称不一致分类：仅当配置包含「物料名称」字段时启用
            if has_name:
                same_name = srow['物料名称'] in code_name_sets.get(code, set())
            else:
                same_name = True
            if has_code and not same_name:
                # 同编码但物料名称完全对不上 → 第四类
                name_mismatch_sub.append(_sub_dict(srow))
                for idx, c in cands_code:
                    if idx not in matched_ins:
                        name_mismatch_ins.append({**_ins_dict(c), '对应送检编码': code})
                        matched_ins.add(idx)
            else:
                # 全键未命中 → 严格核对：
                # 1) 优先寻找匹配字段完全一致的候选 → 转为已检验
                # 2) 否则记录差异原因（重点：数量 / 供应商）
                d = _sub_dict(srow)
                reasons = []
                matched_idx = None
                matched_ins_row = None
                for idx, c in cands_code:
                    if idx in matched_ins:
                        continue
                    # 快速必然性检查：非供应商字段已不同 → 不可能匹配，只需前 2 条差异文本
                    if not _quick_fields_equal(srow, c, pairs):
                        if len(reasons) < 2:
                            reasons.append(_diff_fields(srow, c, cfg, use_norm_supplier=normalize_suppliers, sup_cache=_sup_cache))
                        continue
                    diff = _diff_fields(srow, c, cfg, use_norm_supplier=normalize_suppliers, sup_cache=_sup_cache)
                    if not diff and (not has_date or _date_ge(c[ins_date_field], srow[sub_date_field])):
                        matched_idx = idx
                        matched_ins_row = c
                        break
                    if len(reasons) < 2:
                        if diff:
                            reasons.append(diff)
                        elif has_date:
                            reasons.append(f"质检日期 {_fmt_date(c[ins_date_field]) or '空'} 早于收料日期 {_fmt_date(srow[sub_date_field]) or '空'}")
                if matched_idx is not None:
                    _mr = matched_ins_row or {}
                    d['匹配检验记录'] = (f"{_mr.get('供应商', '') or ''} · {_mr.get('规格型号', '') or ''} · "
                                      f"检验数量 {_fmt_qty(_mr.get('检验数量'))} · "
                                      f"质检日期 {_fmt_date(_mr.get('质检日期'))}")
                    checked.append(d)
                    matched_ins.add(matched_idx)
                else:
                    if reasons:
                        d['未匹配原因'] = '；'.join(reasons[:2])
                    unchecked.append(d)

        if progress and (s_idx + 1) % report_every == 0:
            progress(s_idx + 1, total_sub)

    if progress:
        progress(total_sub, total_sub)

    # 额外检验：未被任何送检匹配的检验记录
    sub_codes = {s['物料编码'] for s in subs} if has_code else set()
    # 统计检验记录的关键字段出现次数，用于识别疑似重复检验单
    dup_counts = {}
    for c in ins_rows:
        _key = _sub_match_key(c) + ((c.get(ins_date_field),) if has_date else ())
        dup_counts[_key] = dup_counts.get(_key, 0) + 1
    extra = []
    for idx, irow in enumerate(ins_rows):
        if idx in matched_ins:
            continue
        note = ''
        if has_code and irow['物料编码'] in sub_codes:
            # 找出同编码的送检记录，尽量给出具体差异（按送检顺序收集前 2 条非空差异即停，结果与原全量收集一致）
            diffs = []
            for s in subs_by_code.get(irow['物料编码'], []):
                d = _diff_fields(s, irow, cfg, use_norm_supplier=normalize_suppliers, sup_cache=_sup_cache)
                if d:
                    diffs.append(d)
                    if len(diffs) >= 2:
                        break
            if diffs:
                note = f"送检清单同编码记录存在差异：{'；'.join(diffs[:2])}"
            else:
                note = '该编码存在于送检清单，但未匹配成功（请核对名称/数量/日期/供应商）'
        # 疑似重复检验单提示（不去重，仅提示人工确认）
        dup_key = _sub_match_key(irow) + ((irow.get(ins_date_field),) if has_date else ())
        if dup_counts.get(dup_key, 0) > 1:
            dup_tip = '该记录与另一条检验记录完全相同，疑似重复检验单，请人工确认'
            note = f"{note}；{dup_tip}" if note else dup_tip
        extra.append({**_ins_dict(irow), '备注': note})

    checked_df = pd.DataFrame(checked, columns=SUBMISSION_COLS + ['采购员', '检验类型', '匹配检验记录'])
    unchecked_df = pd.DataFrame(unchecked, columns=SUBMISSION_COLS + ['采购员', '检验类型', '未匹配原因'])
    extra_df = pd.DataFrame(extra, columns=INSPECTION_COLS + ['采购员', '检验类型', '备注'])
    mismatch_sub_df = pd.DataFrame(name_mismatch_sub, columns=SUBMISSION_COLS + ['采购员', '检验类型'])
    mismatch_ins_df = pd.DataFrame(name_mismatch_ins, columns=INSPECTION_COLS + ['采购员', '检验类型', '对应送检编码'])

    summary = {
        'total_sub': total_sub,
        'checked': len(checked),
        'unchecked': len(unchecked),
        'name_mismatch': len(name_mismatch_sub),
        'extra': len(extra),
    }
    return {
        'checked': checked_df,
        'unchecked': unchecked_df,
        'extra': extra_df,
        'name_mismatch': {'sub': mismatch_sub_df, 'ins': mismatch_ins_df},
        'summary': summary,
    }


# ==================== 导出 ====================

def _to_xlsx_bytes(df, sheet_name='结果', red_rows=None):
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        if red_rows:
            from openpyxl.styles import PatternFill
            fill = PatternFill(start_color='FFF4C4C4', end_color='FFF4C4C4', fill_type='solid')
            ws = writer.sheets[sheet_name]
            for r in red_rows:
                for c in range(1, len(df.columns) + 1):
                    ws.cell(row=r + 2, column=c).fill = fill
    return buf.getvalue()


def export_unchecked(result):
    """导出未检验清单（整表红底高亮）"""
    df = result['unchecked']
    return _to_xlsx_bytes(df, '未检验', red_rows=list(range(len(df))))


def export_all(result):
    """导出全部对比结果（多 sheet，未检验 sheet 红底高亮）"""
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        result['checked'].to_excel(writer, index=False, sheet_name='已检验')
        result['unchecked'].to_excel(writer, index=False, sheet_name='未检验')
        result['extra'].to_excel(writer, index=False, sheet_name='额外检验')
        result['name_mismatch']['sub'].to_excel(writer, index=False, sheet_name='名称不一致-送检')
        result['name_mismatch']['ins'].to_excel(writer, index=False, sheet_name='名称不一致-检验')
        from openpyxl.styles import PatternFill
        fill = PatternFill(start_color='FFF4C4C4', end_color='FFF4C4C4', fill_type='solid')
        ws = writer.sheets['未检验']
        for r in range(2, len(result['unchecked']) + 2):
            for c in range(1, len(result['unchecked'].columns) + 1):
                ws.cell(row=r, column=c).fill = fill
    return buf.getvalue()


# ==================== 检验记录入库（方案四：持久化 + 跨窗口对账） ====================

def parse_inspection_full(uploaded_file, clean_report=None):
    """
    解析检验清单 Excel，保留入库所需完整字段。
    返回 DataFrame（6 核心列 + 扩展列：单据编号/合格数/不合格数/检验结果/质检员/批号/类别/采购员）。

    供应商空值沿用「收料日期填充规则」：若含「单据编号」列，则按单据号分组统一填充
    供应商（以及单据状态/整单关闭状态等 ERP 导出时仅首行有值的列），与送检清单解析行为一致。
    填充明细写入 clean_report（若传入 dict）。
    """
    df = pd.read_excel(uploaded_file, header=None)
    header_row = _find_header_row(df)
    if header_row is None:
        raise ValueError('未找到表头行，请确保包含「物料编码」「物料名称」等列名')
    df.columns = df.iloc[header_row].astype(str).str.strip()
    df = df.iloc[header_row + 1:].reset_index(drop=True)
    df = df.dropna(how='all')

    # ERP 单据格式（含「单据编号」列）：按单据号分组填充供应商/采购员等空值（沿用收料日期填充规则）
    norm = {_normalize_header(c): c for c in df.columns}
    rename_map = {}
    for _key in ('单据编号', '供应商', '整单关闭状态', '采购员'):
        if _key in norm and norm[_key] != _key:
            rename_map[norm[_key]] = _key
    if rename_map:
        df = df.rename(columns=rename_map)
    if '单据编号' in df.columns:
        df, clean_notes = fill_receipt_order(df)
        if isinstance(clean_report, dict):
            clean_report.update(clean_notes)

    df, missing = _map_columns(df, 'inspection')
    if missing:
        raise ValueError(f'缺少必要列: {", ".join(missing)}（请使用标准列名或下载模板）')

    df['供应商'] = df['供应商'].map(_clean_text)
    # 过滤 Excel 底部的「合计」汇总行
    df = df[~df['供应商'].str.contains('合计', na=False)].reset_index(drop=True)
    df['物料编码'] = df['物料编码'].map(_normalize_code)
    df['规格型号'] = df['规格型号'].map(_clean_text)
    df['物料名称'] = df['物料名称'].map(_clean_text)
    df['质检日期'] = df['质检日期'].map(lambda v: _parse_date(v, default=None))
    df['检验数量'] = df['检验数量'].map(_parse_qty)
    df = df[df['物料编码'] != ''].reset_index(drop=True)

    # 提取扩展列（按别名匹配，缺失时置空）
    df_cols = [_normalize_header(c) for c in df.columns]
    for std, aliases in EXTRA_INS_COLS.items():
        if std in df.columns:
            continue
        for i, col in enumerate(df_cols):
            if any(a.lower() == col.lower() for a in aliases):
                df[std] = df[df.columns[i]]
                break
        else:
            df[std] = ''

    for std in ('合格数', '不合格数'):
        if std in df.columns:
            df[std] = df[std].map(_parse_qty)
    for std in ('检验结果', '质检员', '批号', '类别', '单据编号', '采购员'):
        if std in df.columns:
            df[std] = df[std].map(_clean_text)

    return df


def import_inspection_records(df, progress=None, batch_size=1000, source_file='',
                              inspect_type='来料检'):
    """
    幂等入库：检验记录持久化（类型+单据编号+供应商+编码+规格+名称+质检日期+数量 去重）。
    返回 (插入数, 跳过数, 总行数)
    """
    df = df.copy()
    total = len(df)
    rows = []
    for _, row in df.iterrows():
        d = row['质检日期']
        rows.append({
            'doc_no': row.get('单据编号', '') or '',
            'supplier': row['供应商'],
            'material_code': row['物料编码'],
            'spec': row['规格型号'],
            'material_name': row['物料名称'],
            'inspect_date': d.isoformat() if isinstance(d, (date, datetime)) else None,
            'inspect_qty': row['检验数量'],
            'qualified_qty': row.get('合格数'),
            'unqualified_qty': row.get('不合格数'),
            'result': row.get('检验结果', '') or '',
            'inspector': row.get('质检员', '') or '',
            'batch_no': row.get('批号', '') or '',
            'category': row.get('类别', '') or '',
            'purchaser': _clean_text(row.get('采购员', '')),
            'inspect_type': inspect_type,
            'source_file': source_file,
            'dedup_key': make_dedup_key(row, inspect_type, 'inspection'),
        })

    inserted = 0
    if rows:
        for i in range(0, len(rows), batch_size):
            batch = rows[i:i + batch_size]
            inserted += supabase_helper.insert_inspection_records(batch)
            if progress:
                progress(min(i + len(batch), total), total)

    skipped = total - inserted
    return inserted, skipped, total


def inspection_records_to_df(records):
    """数据库检验记录 → 展示/比对用 DataFrame（检验类型 + 6 标准列 + 扩展列 + 入库时间 + id）"""
    rows = [{
        '检验类型': _clean_text(r.get('inspect_type')) or '来料检',
        '供应商': _clean_text(r.get('supplier')),
        '物料编码': _normalize_code(r.get('material_code')),
        '规格型号': _clean_text(r.get('spec')),
        '物料名称': _clean_text(r.get('material_name')),
        '质检日期': _fmt_date(r.get('inspect_date')),
        '检验数量': _parse_qty(r.get('inspect_qty')),
        '单据编号': _clean_text(r.get('doc_no')),
        '合格数': _parse_qty(r.get('qualified_qty')),
        '不合格数': _parse_qty(r.get('unqualified_qty')),
        '检验结果': _clean_text(r.get('result')),
        '质检员': _clean_text(r.get('inspector')),
        '批号': _clean_text(r.get('batch_no')),
        '类别': _clean_text(r.get('category')),
        '采购员': _clean_text(r.get('purchaser')),
        '入库时间': str(r.get('created_at', ''))[:19],
        'id': r.get('id'),
        '去重键': _clean_text(r.get('dedup_key')),
    } for r in records]
    cols = ['检验类型', '供应商', '物料编码', '规格型号', '物料名称', '质检日期', '检验数量',
            '单据编号', '合格数', '不合格数', '检验结果', '质检员', '批号', '类别', '采购员',
            '入库时间', 'id', '去重键']
    return pd.DataFrame(rows, columns=cols)
